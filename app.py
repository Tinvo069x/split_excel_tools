import streamlit as st
import pandas as pd
import io, zipfile
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook, Workbook
from copy import copy

def copy_row_with_format(src_ws, tgt_ws, src_row, tgt_row):
    """Copy nguyên giá trị, công thức, format từ 1 row sang row mới"""
    for col, cell in enumerate(src_ws[src_row], start=1):
        new_cell = tgt_ws.cell(row=tgt_row, column=col, value=cell.value)
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)

def split_excel_with_format(input_file, sheet_name, header_row, selected_cols):
    # Dùng pandas để group dữ liệu (nhanh)
    df = pd.read_excel(input_file, sheet_name=sheet_name, header=header_row-1, dtype=str)
    df = df.fillna("")

    # Load workbook bằng openpyxl để copy format
    wb = load_workbook(input_file, data_only=False)  # data_only=False để giữ công thức
    ws = wb[sheet_name]

    bad_words = ["(All)", "Sum of", "Supplier", "Invoice", "Shipmode"]

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zf:
        for keys, group in df.groupby(selected_cols):
            if isinstance(keys, tuple):
                split_key = "-".join([str(k).strip() for k in keys])
            else:
                split_key = str(keys).strip()

            if not split_key or any(bw.lower() in split_key.lower() for bw in bad_words):
                continue

            for ch in r'\/:*?"<>|':
                split_key = split_key.replace(ch, "_")

            # Tạo workbook mới
            new_wb = Workbook()
            new_ws = new_wb.active

            # Copy header block
            for r in range(1, header_row+1):
                copy_row_with_format(ws, new_ws, r, r)

            # Copy từng dòng dữ liệu tương ứng
            paste_row = header_row + 1
            for _, row in group.iterrows():
                excel_row_idx = row.name + header_row + 1  # index pandas + offset
                copy_row_with_format(ws, new_ws, excel_row_idx, paste_row)
                paste_row += 1

            # Lưu vào memory
            output = io.BytesIO()
            file_name = f"{split_key}-{datetime.today().strftime('%Y%m%d')}.xlsx"
            new_wb.save(output)
            zf.writestr(file_name, output.getvalue())

    return zip_buffer

# ===================== Streamlit UI =====================
st.title("📊 Split Excel xlsx, xlsm-Tins")

uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx","xlsm"])

if uploaded_file:
    xls = pd.ExcelFile(uploaded_file, engine="openpyxl")
    sheet_name = st.selectbox("Chọn sheet:", xls.sheet_names)
    header_row = st.number_input("Chọn dòng header", min_value=1, value=1, step=1)

    if st.button("🔍 Xem trước"):
        df_preview = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=header_row-1, nrows=10)
        st.dataframe(df_preview)

    df = pd.read_excel(uploaded_file, sheet_name=sheet_name, header=header_row-1, dtype=str)
    selected_cols = st.multiselect("Chọn cột để split:", df.columns.tolist())

    if st.button("🚀 Split Now"):
        if not selected_cols:
            st.warning("⚠️ Vui lòng chọn ít nhất 1 cột")
        else:
            zip_buffer = split_excel_with_format(uploaded_file, sheet_name, header_row, selected_cols)
            st.success("✅ Đã tách file, giữ nguyên công thức và định dạng!")

            st.download_button(
                label="📥 Download ZIP",
                data=zip_buffer.getvalue(),
                file_name=f"SplitResult-{datetime.today().strftime('%Y%m%d')}.zip",
                mime="application/zip"
            )
